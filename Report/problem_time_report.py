import requests
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.chart import PieChart, Reference

# Dynatrace API bilgileri
API_TOKEN = "--"
API_URL = "--"
DYNATRACE_BASE_URL = "--"

# Zaman aralığı (son 1 gün)
end_time = datetime.utcnow()
start_time = end_time - timedelta(days=1)

# Mikro saniyeleri kaldırarak ISO 8601 formatında string oluştur
start_time_iso = start_time.replace(microsecond=0).isoformat() + "Z"
end_time_iso = end_time.replace(microsecond=0).isoformat() + "Z"

# İlk API isteği için parametreler
params = {
    "from": start_time_iso,
    "to": end_time_iso,
    "pageSize": 10,  # Maksimum değer 10
    "fields": "statusDetails,entityDetails"
}

# API başlıkları
headers = {
    "Authorization": f"Api-Token {API_TOKEN}"
}

# Sayfalama için liste
all_problems = []

while True:
    # API isteği
    response = requests.get(API_URL, headers=headers, params=params)

    # Yanıtı kontrol et
    if response.status_code != 200:
        print("API isteği başarısız:", response.status_code, response.text)
        exit()

    # JSON verilerini al
    data = response.json()

    # Gelen problemleri listeye ekle
    all_problems.extend(data.get("problems", []))

    # Sonraki sayfa anahtarını kontrol et
    next_page_key = data.get("nextPageKey")
    if not next_page_key:
        break

    # Bir sonraki sayfa için sadece nextPageKey parametresini gönder
    params = {
        "nextPageKey": next_page_key
    }

# Problemleri analiz et
problem_details = []

for problem in all_problems:
    try:
        # Zaman damgası milisaniyeden saniyeye dönüştürülüyor
        open_time = datetime.fromtimestamp(problem["startTime"] / 1000)

        # endTime değerini kontrol et
        if problem.get("endTime") and problem["endTime"] > 0:
            close_time = datetime.fromtimestamp(problem["endTime"] / 1000)
        else:
            close_time = end_time

        # Açık kalma süresi hesaplama
        duration_minutes = (close_time - open_time).total_seconds() / 60

        # Kategorilere ayırma
        if duration_minutes == 0:
            category = "0 dk"
        elif 1 <= duration_minutes <= 5:
            category = "1-5 dk"
        elif 5 < duration_minutes <= 10:
            category = "5-10 dk"
        elif 10 < duration_minutes <= 30:
            category = "10-30 dk"
        else:
            category = "30 dk üzeri"

        # Etkilenen entity bilgileri
        affected_entities = ", ".join(entity.get("name", "Unknown") for entity in problem.get("affectedEntities", []))

        # Problem URL'si
        problem_url = f"{DYNATRACE_BASE_URL}/#problems/problemdetails;pid={problem['problemId']}"

        problem_details.append({
            "Problem ID": problem["problemId"],
            "Başlık": problem["title"],
            "Açık Kalma Süresi (dk)": round(duration_minutes, 2),
            "Kategori": category,
            "Etkilenen Entities": affected_entities,
            "Problem URL": problem_url
        })
    except Exception as e:
        print(f"Problem işlenirken hata oluştu: {e}")
        continue

# Veriyi Excel'e yaz
df = pd.DataFrame(problem_details)

excel_file = "problem_report.xlsx"
with pd.ExcelWriter(excel_file) as writer:
    # Her kategori için ayrı bir çalışma sayfası oluştur
    for category in df["Kategori"].unique():
        category_df = df[df["Kategori"] == category]
        category_df.to_excel(writer, sheet_name=category, index=False)
    # Genel özet için bir sayfa ekleyelim
    summary_df = df["Kategori"].value_counts().reset_index()
    summary_df.columns = ["Kategori", "Toplam Problem Sayısı"]
    summary_df.to_excel(writer, sheet_name="Özet", index=False)

# Pasta grafiği ekle
wb = load_workbook(excel_file)
ws = wb["Özet"]

# Pasta grafiği oluştur
chart = PieChart()
labels = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
chart.add_data(data, titles_from_data=True)
chart.set_categories(labels)
chart.title = "Problem Kategorileri"

# Grafiği ekle
ws.add_chart(chart, "E5")

# Excel dosyasını kaydet
wb.save(excel_file)

print(f"Rapor '{excel_file}' dosyasına kaydedildi.")