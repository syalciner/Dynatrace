import requests 
import pandas as pd
from datetime import datetime, timedelta
from collections import Counter
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
import matplotlib.pyplot as plt

# Dynatrace API bilgileri
dynatrace_url = "--/api/v2/problems"
api_token = "--"
notifications_url = "--/api/v1/notifications"
settings_url = "--/api/v2/settings/objects"

# 1 haftalık zaman aralığı
end_time = datetime.utcnow()
start_time = end_time - timedelta(days=7)
start_time_iso = start_time.strftime("%Y-%m-%dT%H:%M:%SZ")
end_time_iso = end_time.strftime("%Y-%m-%dT%H:%M:%SZ")

# API çağrısı için headers
headers = {
    "Authorization": f"Api-Token {api_token}"
}

# API çağrısı için başlangıç parametreleri
initial_params = {
    "from": start_time_iso,
    "to": end_time_iso,
    "pageSize": 100,
}

def convert_epoch_to_datetime(epoch_time):
    if epoch_time:
        return datetime.utcfromtimestamp(epoch_time / 1000).strftime("%Y-%m-%d %H:%M:%S")
    return "N/A"

def get_problem_details(problem_id):
    problem_detail_url = f"{dynatrace_url}/{problem_id}"
    response = requests.get(problem_detail_url, headers=headers)
    response.raise_for_status()
    return response.json()

def process_list(data_list, key="name"):
    if isinstance(data_list, list) and data_list:
        return [item.get(key, "Unknown") if isinstance(item, dict) else str(item) for item in data_list]
    return []

def fetch_all_problems():
    problems = []
    next_page_key = None
    params = initial_params.copy()

    while True:
        if next_page_key:
            params = {"nextPageKey": next_page_key}

        response = requests.get(dynatrace_url, headers=headers, params=params)
        response.raise_for_status()
        data = response.json()

        if "problems" in data:
            problems.extend(data["problems"])

        next_page_key = data.get("nextPageKey")
        if not next_page_key:
            break

    return problems

def fetch_alerting_profiles():
    # Settings endpoint'i kullanarak builtin:alerting.profile objelerini çeker
    params = {
        "schemaIds": "builtin:alerting.profile",
        "pageSize": 200
    }
    response = requests.get(settings_url, headers=headers, params=params)
    response.raise_for_status()
    data = response.json()

    profile_id_map = {}
    items = data.get("items", [])
    for item in items:
        object_id = item.get("objectId")
        value = item.get("value", {})
        profile_name = value.get("name")
        if profile_name and object_id:
            profile_id_map[profile_name] = object_id

    return profile_id_map

#####################
# İKİNCİ KOD ENTEGRASYONU
#####################

DYNATRACE_BASE_URL = "--/api/v2"
DYNATRACE_API_TOKEN = api_token

NOTIF_HEADERS = {
    "Authorization": f"Api-Token {DYNATRACE_API_TOKEN}",
    "Content-Type": "application/json"
}

def list_email_notifications():
    endpoint = f"{DYNATRACE_BASE_URL}/settings/objects"
    params = {
        "schemaIds": "builtin:problem.notifications",
        "pageSize": 200
    }

    email_notifications = []
    next_page_key = None

    while True:
        if next_page_key:
            params['nextPageKey'] = next_page_key

        response = requests.get(endpoint, headers=NOTIF_HEADERS, params=params)
        if response.status_code != 200:
            print(f"Error fetching notifications: {response.status_code}, {response.text}")
            return []

        data = response.json()
        for item in data.get("items", []):
            # type EMAIL olanları çek
            if item.get("value", {}).get("type") == "EMAIL":
                email_notifications.append(item)

        next_page_key = data.get("nextPageKey")
        if not next_page_key:
            break

    return email_notifications

def build_alerting_profile_recipients_map(email_notifications):
    # Bu fonksiyon alertingProfile object_id -> recipients bilgisi şeklinde map oluşturur.
    profile_recipients_map = {}
    for notification in email_notifications:
        value = notification.get('value', {})
        alerting_profile_id = value.get('alertingProfile', '')
        email_info = value.get('emailNotification', {})

        # Receipient bilgilerinin birleştirilmesi
        recipients_list = []
        if email_info.get('recipients'):
            recipients_list.extend(email_info['recipients'])
        if email_info.get('ccRecipients'):
            recipients_list.extend([f"CC: {r}" for r in email_info['ccRecipients']])
        if email_info.get('bccRecipients'):
            recipients_list.extend([f"BCC: {r}" for r in email_info['bccRecipients']])

        recipients_str = ", ".join(recipients_list)
        # Aynı alerting profile id'ye birden fazla notification tanımı olabilir.
        # Hepsini birleştiriyoruz.
        if alerting_profile_id:
            if alerting_profile_id in profile_recipients_map:
                # Aynı profile için birden fazla tanım varsa birleştir
                existing = profile_recipients_map[alerting_profile_id]
                if recipients_str not in existing:
                    profile_recipients_map[alerting_profile_id] = existing + "; " + recipients_str
            else:
                profile_recipients_map[alerting_profile_id] = recipients_str

    return profile_recipients_map

#####################
# ANA İŞLEM
#####################

# Verileri çek
problems_data = fetch_all_problems()
profile_id_map = fetch_alerting_profiles()

problems = []

alerting_profile_counts = Counter()
entity_counts = Counter()
entity_alerting_profiles = {}

for problem in problems_data:
    problem_id = problem.get("problemId")
    details = get_problem_details(problem_id)
    
    management_zones = process_list(details.get("managementZones", []))
    problem_filters = details.get("problemFilters", [])
    alerting_profiles = process_list(problem_filters, key="name")

    # "Default" profilini çıkar
    alerting_profiles = [profile for profile in alerting_profiles if profile != "Default"]

    # Alerting Profile ID'lerini eşleştir
    alerting_profile_ids = [profile_id_map.get(profile, "Unknown") for profile in alerting_profiles]

    # Sayımları artır
    for profile in alerting_profiles:
        alerting_profile_counts[profile] += 1

    affected_entities = problem.get("affectedEntities", [])
    for entity in affected_entities:
        entity_name = entity.get("name", "Unknown")
        entity_counts[entity_name] += 1
        entity_alerting_profiles[entity_name] = ", ".join(alerting_profiles)

    problems.append({
        "Problem ID": problem_id,
        "Title": problem.get("title"),
        "Status": problem.get("status"),
        "Severity Level": problem.get("severityLevel"),
        "Impact Level": problem.get("impactLevel"),
        "Start Time": convert_epoch_to_datetime(problem.get("startTime")),
        "End Time": convert_epoch_to_datetime(problem.get("endTime")),
        "Management Zones": ", ".join(management_zones),
        "Alerting Profiles": ", ".join(alerting_profiles),
        "Alerting Profile IDs": ", ".join(alerting_profile_ids),
        "Affected Entities": ", ".join([entity.get("name", "Unknown") for entity in affected_entities]),
    })

# DataFrame'ler
df = pd.DataFrame(problems)

summary_data = []
for profile, count in alerting_profile_counts.items():
    summary_data.append({
        "Alerting Profile": profile, 
        "Alerting Profile ID": profile_id_map.get(profile, "Unknown"),
        "Problem Count": count
    })
summary_df = pd.DataFrame(summary_data).sort_values(by="Problem Count", ascending=False)

entity_summary_data = []
for entity, count in entity_counts.items():
    profiles_str = entity_alerting_profiles.get(entity, "")
    profile_ids = [profile_id_map.get(p.strip(), "Unknown") for p in profiles_str.split(",") if p.strip()]
    entity_summary_data.append({
        "Entity": entity, 
        "Problem Count": count, 
        "Alerting Profiles": profiles_str,
        "Alerting Profile IDs": ", ".join(profile_ids)
    })
entity_summary_df = pd.DataFrame(entity_summary_data).sort_values(by="Problem Count", ascending=False)

# Grafikler
top_15_alerting_profiles = summary_df.head(15)
top_15_entities = entity_summary_df.head(15)

# Pie Chart - İlk 15 Alerting Profiles
plt.figure(figsize=(8, 8))
plt.pie(top_15_alerting_profiles["Problem Count"], labels=top_15_alerting_profiles["Alerting Profile"], autopct='%1.1f%%', startangle=140)
plt.title("Top 15 Alerting Profiles")
alerting_profile_chart = "alerting_profile_pie.png"
plt.savefig(alerting_profile_chart)
plt.close()

# Pie Chart - İlk 15 Entities
plt.figure(figsize=(8, 8))
plt.pie(top_15_entities["Problem Count"], labels=top_15_entities["Entity"], autopct='%1.1f%%', startangle=140)
plt.title("Top 15 Entities")
entity_chart = "entity_pie.png"
plt.savefig(entity_chart)
plt.close()

#####################
# İKİNCİ KOD ENTEGRASYONU: RECIPIENTS
#####################
email_notifications = list_email_notifications()
profile_recipients_map = build_alerting_profile_recipients_map(email_notifications)

# summary_df içerisine Recipients kolonu ekle
summary_df["Recipients"] = summary_df["Alerting Profile ID"].map(profile_recipients_map).fillna("")

# Excel'e yazdır
excel_filename = "Dynatrace_Problems_Report.xlsx"
with pd.ExcelWriter(excel_filename, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Problems", index=False)
    summary_df.to_excel(writer, sheet_name="Alerting Profiles", index=False)
    entity_summary_df.to_excel(writer, sheet_name="Entities", index=False)

# Grafikleri ekle
wb = load_workbook(excel_filename)
ws = wb.create_sheet(title="Charts")

img = Image(alerting_profile_chart)
ws.add_image(img, "A1")

img = Image(entity_chart)
ws.add_image(img, "A20")

wb.save(excel_filename)

print(f"Rapor başarıyla {excel_filename} dosyasına kaydedildi!")